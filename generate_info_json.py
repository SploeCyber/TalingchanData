import os
import json
from collections import OrderedDict
from tkinter import Tk, filedialog, simpledialog
from openpyxl import load_workbook
from PIL import Image
from io import BytesIO

Tk().withdraw()
xlsx_path = filedialog.askopenfilename(
    title="‡πÄ‡∏•‡∏∑‡∏≠‡∏Å‡πÑ‡∏ü‡∏•‡πå Excel (.xlsx)",
    filetypes=[("Excel files", "*.xlsx")]
)
if not xlsx_path:
    exit()

output_folder = simpledialog.askstring("Output Folder", "‡∏ï‡∏±‡πâ‡∏á‡∏ä‡∏∑‡πà‡∏≠‡πÇ‡∏ü‡∏•‡πÄ‡∏î‡∏≠‡∏£‡πå‡∏™‡∏≥‡∏´‡∏£‡∏±‡∏ö‡∏£‡∏π‡∏õ‡πÅ‡∏•‡∏∞ JSON:")
if not output_folder:
    output_folder = "images"
os.makedirs(output_folder, exist_ok=True)

wb = load_workbook(xlsx_path, data_only=True)
ws = wb.active

images_data = {}
for image in ws._images:
    img_bytes = image._data()
    if isinstance(img_bytes, bytes):
        img = Image.open(BytesIO(img_bytes))
        if image.anchor._from:
            row = image.anchor._from.row + 1
            images_data[row] = img

def get_merged_value(cell):
    for merged in ws.merged_cells.ranges:
        if cell.coordinate in merged:
            top_left = ws.cell(merged.min_row, merged.min_col)
            return top_left.value
    return cell.value

cards = []
row = 1
while True:
    if ws.cell(row=row, column=2).value is None and row not in images_data:
        break
    card_temp = {}
    first_key = ws.cell(row=row, column=3).value
    if first_key:
        keys = [ws.cell(row=row, column=c).value for c in range(3, 15)]
        vals = [ws.cell(row=row+1, column=c).value for c in range(3, 15)]
        for k, v in zip(keys, vals):
            if k:
                card_temp[str(k)] = v
    details = {}
    for r in range(row+2, row+5):
        key_cell = ws.cell(row=r, column=3)
        val_cell = ws.cell(row=r, column=4)
        key_val = get_merged_value(key_cell)
        val_val = get_merged_value(val_cell)
        if key_val:
            details[str(key_val)] = val_val
    if details:
        card_temp["Details"] = details
    bg_color = ws.cell(row=row+1, column=9).fill.fgColor.rgb
    if bg_color and bg_color != "00000000":
        card_temp["Color"] = f"#{bg_color[-6:]}"
    else:
        card_temp["Color"] = None
    img_obj = images_data.get(row)
    if img_obj:
        print_code = str(card_temp.get("Print", f"card_{len(cards)+1}")).strip()
        rarity = str(card_temp.get("Rare", "")).strip()
        name_part = f"{print_code} ({rarity})" if rarity else print_code
        safe_name = name_part.replace("/", "-").replace("\\", "-")
        img_path = os.path.join(output_folder, f"{safe_name}.png")
        img_obj.save(img_path)
        card_temp["ImagePath"] = img_path
    else:
        card_temp["ImagePath"] = None

    if str(card_temp.get("Type", "")).strip().lower() == "life":
        card_temp = {k: v for k, v in card_temp.items() if v not in (None, "", " ")}

    card = OrderedDict()
    card["ImagePath"] = card_temp.pop("ImagePath", None)
    for k, v in card_temp.items():
        card[k] = v
    cards.append(card)
    row += 5

json_path = os.path.join(output_folder, "info.json")
with open(json_path, "w", encoding="utf-8") as f:
    json.dump(cards, f, ensure_ascii=False, indent=2)

print(f"‚úÖ ‡πÅ‡∏õ‡∏•‡∏á‡πÄ‡∏™‡∏£‡πá‡∏à‡πÅ‡∏•‡πâ‡∏ß: {json_path}")
print(f"üñºÔ∏è ‡∏£‡∏π‡∏õ‡∏†‡∏≤‡∏û‡πÅ‡∏•‡∏∞ JSON ‡∏ñ‡∏π‡∏Å‡∏ö‡∏±‡∏ô‡∏ó‡∏∂‡∏Å‡πÉ‡∏ô‡πÇ‡∏ü‡∏•‡πÄ‡∏î‡∏≠‡∏£‡πå: {output_folder}")